from __future__ import annotations

import csv
import re
from io import BytesIO, StringIO
from typing import Any

from app.domain.kc_career_path import dump_career_path, parse_career_path
from app.domain.kc_format import cell_value_to_str, normalize_kc_field_value

DEFAULT_DEPARTMENT = "Служба поддержки"

# Заголовки как в исходной таблице (порядок колонок).
KC_TEMPLATE_HEADERS: list[str] = [
    "Л",
    "К",
    "Г",
    "ФИО",
    "Должность",
    "Почта новая",
    "Телефон",
    "Адрес проживания",
    "Имя пользователя в Telegram",
    "Номер счета",
    "ID Telegram",
    "Дата Рождения",
    "Первый день работы",
    "Допуск (дата)",
    "Дата ухода/перехода",
]

# Опционально для группировки на экране (не в исходном списке, но полезно).
KC_OPTIONAL_HEADER_DEPARTMENT = "Отдел"
KC_OPTIONAL_HEADER_SUBDIVISION = "Подраздел"

KC_EXPORT_HEADERS: list[str] = KC_TEMPLATE_HEADERS + [
    KC_OPTIONAL_HEADER_DEPARTMENT,
    KC_OPTIONAL_HEADER_SUBDIVISION,
]

KC_EXAMPLE_ROW: list[str] = [
    "1",
    "Пари",
    "Ростов-на-Дону",
    "Иванов Иван Иванович",
    "Оператор 1 линии",
    "ivanov@pari.ru",
    "+7 900 000-00-00",
    "г. Ростов-на-Дону, ул. Примерная, 1",
    "@ivanov",
    "40817810****0000",
    "123456789",
    "01.01.1995",
    "08.05.2023",
    "10.05.2023",
    "",
]

KC_DB_FIELD_LIMITS: dict[str, int] = {
    "line": 32,
}

_FIELD_ALIASES: dict[str, list[str]] = {
    "line": ["л", "линия", "line"],
    "company": ["к", "компания", "company"],
    "city": ["г", "город", "city"],
    "full_name": ["фио", "ф.и.о.", "fullname", "full name"],
    "position": ["должность", "position"],
    "grade_new": ["грейд new", "грейд", "grade new", "grade"],
    "email_new": ["почта новая", "почта", "email", "e-mail"],
    "phone": ["телефон", "phone"],
    "residence_address": ["адрес проживания", "адрес"],
    "telegram_username": ["имя пользователя в telegram", "telegram", "телеграм"],
    "account_number": ["номер счета", "счет", "счёт"],
    "account_number_extra": ["дополнительный счёт", "дополнительный счет", "доп. счёт", "доп. счет"],
    "telegram_id": ["id telegram", "telegram id"],
    "birth_date": ["дата рождения", "др"],
    "first_work_day": ["первый день работы", "дата выхода"],
    "access_date": ["допуск (дата)", "допуск"],
    "leave_or_transfer_date": ["дата ухода/перехода", "дата ухода", "уход"],
    "department": ["отдел", "department"],
    "subdivision": ["подраздел", "разветвление", "subdivision"],
}


def _norm_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("\n", " ").replace("\r", " ")
    text = re.sub(r"\s+", " ", text)
    return text


def _map_header_to_field(header: str) -> str | None:
    h = _norm_header(header)
    if not h:
        return None
    for field, aliases in _FIELD_ALIASES.items():
        if h in aliases:
            return field
    return None


def _cell_str(value: Any, field: str | None = None) -> str:
    return cell_value_to_str(value, field)


def _finalize_import_row(raw: dict[str, str]) -> dict[str, str]:
    return {key: normalize_kc_field_value(key, value) for key, value in raw.items()}


def enforce_import_row_limits(row: dict[str, str], row_number: int) -> None:
    for field, limit in KC_DB_FIELD_LIMITS.items():
        value = row.get(field, "")
        if len(value) > limit:
            label = "Линия" if field == "line" else field
            raise ValueError(
                f"Строка {row_number}, «{row.get('full_name', '')}»: "
                f"поле «{label}» слишком длинное ({len(value)} симв., макс. {limit}). "
                f"Значение: «{value[:48]}{'…' if len(value) > 48 else ''}»"
            )


def row_dict_from_values(row_values: tuple[Any, ...] | list[Any], header_map: dict[int, str]) -> dict[str, str]:
    raw: dict[str, str] = {}
    for idx, field in header_map.items():
        if idx < len(row_values):
            raw[field] = _cell_str(row_values[idx], field)
    full_name = raw.get("full_name", "")
    if not full_name:
        return {}
    department = raw.get("department", "").strip() or DEFAULT_DEPARTMENT
    email = raw.get("email_new", "")
    leave_date = raw.get("leave_or_transfer_date", "")
    row = {
        "department": department,
        "subdivision": raw.get("subdivision", ""),
        "line": raw.get("line", ""),
        "company": raw.get("company", ""),
        "city": raw.get("city", ""),
        "full_name": full_name,
        "position": raw.get("position", ""),
        "grade_new": raw.get("grade_new", ""),
        "email_new": email,
        "phone": raw.get("phone", ""),
        "skype": "",
        "residence_address": raw.get("residence_address", ""),
        "telegram_username": raw.get("telegram_username", ""),
        "account_number": raw.get("account_number", ""),
        "account_number_extra": raw.get("account_number_extra", ""),
        "telegram_id": raw.get("telegram_id", ""),
        "birth_date": raw.get("birth_date", ""),
        "first_work_day": raw.get("first_work_day", ""),
        "access_date": raw.get("access_date", ""),
        "leave_or_transfer_date": leave_date,
        "career_path": dump_career_path(parse_career_path("", leave_date)),
        "photo_url": "",
    }
    return _finalize_import_row(row)


def employee_to_export_row(row: Any) -> list[str]:
    """Строка Excel из модели KcEmployee (openpyxl-лист «Актуальный список»)."""
    from app.domain.kc_career_path import parse_career_path

    steps = parse_career_path(getattr(row, "career_path", "[]") or "[]", row.leave_or_transfer_date)
    leave_date = steps[-1].get("date", "") if steps else (row.leave_or_transfer_date or "")
    return [
        row.line or "",
        row.company or "",
        row.city or "",
        row.full_name or "",
        row.position or "",
        row.email_new or "",
        row.phone or "",
        row.residence_address or "",
        row.telegram_username or "",
        row.account_number or "",
        row.telegram_id or "",
        row.birth_date or "",
        row.first_work_day or "",
        row.access_date or "",
        leave_date,
        row.department or "",
        row.subdivision or "",
    ]


def build_template_xlsx(actual_rows: list[list[str]] | None = None) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws_tpl = wb.active
    ws_tpl.title = "Шаблон"
    headers = KC_TEMPLATE_HEADERS + [KC_OPTIONAL_HEADER_DEPARTMENT]
    ws_tpl.append(headers)
    example = KC_EXAMPLE_ROW + ["Служба поддержки"]
    ws_tpl.append(example)
    for cell in ws_tpl[1]:
        cell.font = Font(bold=True)
    ws_tpl.freeze_panes = "A2"

    ws_list = wb.create_sheet("Актуальный список")
    ws_list.append(KC_EXPORT_HEADERS)
    for cell in ws_list[1]:
        cell.font = Font(bold=True)
    for row in actual_rows or []:
        ws_list.append(row)
    ws_list.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _parse_xlsx_sheet(ws) -> list[dict[str, str]]:
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return []

    header_map: dict[int, str] = {}
    for idx, cell in enumerate(header_row):
        field = _map_header_to_field(_cell_str(cell))
        if field:
            header_map[idx] = field

    if "full_name" not in header_map.values():
        raise ValueError("В файле нет колонки «ФИО»")

    result: list[dict[str, str]] = []
    for row in rows_iter:
        if not any(c is not None and str(c).strip() for c in row):
            continue
        parsed = row_dict_from_values(row, header_map)
        if parsed:
            result.append(parsed)
    return result


def parse_xlsx(content: bytes) -> list[dict[str, str]]:
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    names = list(wb.sheetnames)
    preferred_names = {
        "актуальный список",
        "шаблон",
        "sheet1",
        "лист1",
    }
    ordered = [n for n in names if n.strip().lower() in preferred_names]
    ordered.extend(n for n in names if n not in ordered)

    last_error: ValueError | None = None
    for sheet_name in ordered:
        try:
            rows = _parse_xlsx_sheet(wb[sheet_name])
        except ValueError as exc:
            last_error = exc
            continue
        if rows:
            return rows

    if last_error:
        raise last_error
    return []


def parse_csv(content: bytes) -> list[dict[str, str]]:
    text = content.decode("utf-8-sig", errors="replace")
    if ";" in text.split("\n", 1)[0] and "\t" not in text.split("\n", 1)[0]:
        delimiter = ";"
    else:
        delimiter = ","
    reader = csv.reader(StringIO(text), delimiter=delimiter)
    try:
        header_row = next(reader)
    except StopIteration:
        return []

    header_map: dict[int, str] = {}
    for idx, cell in enumerate(header_row):
        field = _map_header_to_field(cell)
        if field:
            header_map[idx] = field

    if "full_name" not in header_map.values():
        raise ValueError("В файле нет колонки «ФИО»")

    result: list[dict[str, str]] = []
    for row in reader:
        if not any(str(c).strip() for c in row):
            continue
        parsed = row_dict_from_values(tuple(row), header_map)
        if parsed:
            result.append(parsed)
    return result


def parse_import_file(content: bytes, filename: str) -> list[dict[str, str]]:
    name = (filename or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return parse_xlsx(content)
    if name.endswith(".csv"):
        return parse_csv(content)
    # Попытка определить по содержимому
    if content[:2] == b"PK":
        return parse_xlsx(content)
    return parse_csv(content)
