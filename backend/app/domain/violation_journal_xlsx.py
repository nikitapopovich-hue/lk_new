from __future__ import annotations

from datetime import datetime
from io import BytesIO

from app.domain.violation_journal import parse_violation_date


def _cell_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%d.%m.%Y")
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def _penalty_kind(text: str) -> str:
    t = text.lower()
    if "штраф" in t or t in ("ш", "fine"):
        return "fine"
    return "warning"


def _penalty_label(kind: str) -> str:
    return "Штраф" if kind == "fine" else "Предупреждение"


def _has_explanation(raw: str) -> bool:
    return raw.lower() in ("да", "yes", "1", "true", "+", "y")


def parse_violation_xlsx(content: bytes) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(content), data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("В файле нет листа")

    header = [_cell_str(c.value) for c in ws[1][:9]]
    if not header or "дата" not in (header[0] or "").lower():
        raise ValueError('Ожидается первая строка с заголовками: «Дата», «Ф.И.О.», …')

    rows: list[dict] = []
    errors: list[str] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not any(row):
            continue
        cells = list(row[:9])
        while len(cells) < 9:
            cells.append(None)
        date_s = _cell_str(cells[0])
        fio = _cell_str(cells[1])
        if not date_s and not fio:
            continue
        if not parse_violation_date(date_s):
            errors.append(f"Строка {row_idx}: некорректная дата «{date_s}»")
            continue
        if not fio:
            errors.append(f"Строка {row_idx}: не указано Ф.И.О.")
            continue
        pen_raw = _cell_str(cells[5])
        pk = _penalty_kind(pen_raw)
        expl_raw = _cell_str(cells[6])
        try:
            fine = float(cells[7]) if cells[7] is not None else 0.0
        except (TypeError, ValueError):
            fine = 0.0
        rows.append(
            {
                "date": date_s,
                "employeeName": fio,
                "recordedBy": _cell_str(cells[2]),
                "groupName": _cell_str(cells[3]),
                "violationType": _cell_str(cells[4]),
                "penaltyKind": pk,
                "hasExplanation": _has_explanation(expl_raw),
                "fineAmount": fine if pk == "fine" else 0,
                "comment": _cell_str(cells[8]),
            }
        )

    if errors and not rows:
        raise ValueError("; ".join(errors[:5]))
    return rows


def build_violation_xlsx(entries: list[dict]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Журнал"

    headers = [
        "Дата",
        "Ф.И.О.",
        "Зафиксировано",
        "Группа",
        "Тип нарушения",
        "Предупреждение / Штраф",
        "Объяснительная",
        "Сумма",
        "Комментарий",
    ]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0A1028")
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, entry in enumerate(entries, start=2):
        pk = entry.get("penaltyKind") or "warning"
        ws.cell(row=row_idx, column=1, value=entry.get("date") or "")
        ws.cell(row=row_idx, column=2, value=entry.get("employeeName") or "")
        ws.cell(row=row_idx, column=3, value=entry.get("recordedBy") or "")
        ws.cell(row=row_idx, column=4, value=entry.get("groupName") or "")
        ws.cell(row=row_idx, column=5, value=entry.get("violationType") or "")
        ws.cell(row=row_idx, column=6, value=_penalty_label(pk))
        ws.cell(row=row_idx, column=7, value="Да" if entry.get("hasExplanation") else "—")
        fine = entry.get("fineAmount") or 0
        ws.cell(row=row_idx, column=8, value=fine if pk == "fine" else "—")
        ws.cell(row=row_idx, column=9, value=entry.get("comment") or "")

    from openpyxl.utils import get_column_letter

    widths = [12, 28, 22, 14, 36, 22, 14, 10, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
