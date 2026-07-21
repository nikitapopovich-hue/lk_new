"""Формирование Excel-отчёта Тригеров РА."""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# Фирменные цвета PARI
_HEADER_FILL = PatternFill("solid", fgColor="0A1028")
_HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
_TITLE_FONT = Font(bold=True, color="0A1028", name="Calibri", size=14)
_SUBTITLE_FONT = Font(color="555555", name="Calibri", size=10)
_SECTION_FILL = PatternFill("solid", fgColor="00C7B1")
_SECTION_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
_RISK_FILL = PatternFill("solid", fgColor="FFEBE9")
_ALT_FILL = PatternFill("solid", fgColor="F3F6FB")
_GOOD_FONT = Font(color="1A7F37", bold=True, name="Calibri", size=10)
_BAD_FONT = Font(color="CF222E", bold=True, name="Calibri", size=10)
_NEUTRAL_FONT = Font(color="57606A", name="Calibri", size=10)
_BODY_FONT = Font(name="Calibri", size=10)
_THIN = Border(
    left=Side(style="thin", color="D0D7DE"),
    right=Side(style="thin", color="D0D7DE"),
    top=Side(style="thin", color="D0D7DE"),
    bottom=Side(style="thin", color="D0D7DE"),
)


def _auto_width(ws: Worksheet, min_w: float = 10, max_w: float = 42) -> None:
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        length = 0
        for cell in col_cells:
            if cell.value is None:
                continue
            length = max(length, len(str(cell.value)))
        ws.column_dimensions[letter].width = max(min_w, min(max_w, length + 2.5))


def _write_title(ws: Worksheet, row: int, title: str, subtitle: str = "") -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = _TITLE_FONT
    cell.alignment = Alignment(vertical="center")
    row += 1
    if subtitle:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        sub = ws.cell(row=row, column=1, value=subtitle)
        sub.font = _SUBTITLE_FONT
        row += 1
    return row + 1


def _write_headers(ws: Worksheet, row: int, headers: list[str]) -> int:
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=title)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN
    ws.row_dimensions[row].height = 28
    return row + 1


def _style_body_cell(cell, *, alt: bool = False, risk: bool = False) -> None:
    cell.font = _BODY_FONT
    cell.border = _THIN
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    if risk:
        cell.fill = _RISK_FILL
    elif alt:
        cell.fill = _ALT_FILL


def _trend_font(sentiment: str) -> Font:
    if sentiment == "good":
        return _GOOD_FONT
    if sentiment == "bad":
        return _BAD_FONT
    return _NEUTRAL_FONT


def _pct(v: Any) -> str | float:
    if v is None:
        return "—"
    try:
        return round(float(v), 1)
    except (TypeError, ValueError):
        return "—"


def _sheet_operators(wb: Workbook, data: dict[str, Any]) -> None:
    ws = wb.active
    ws.title = "Операторы СП+VIP"
    period_days = data.get("periodDays") or ""
    period = data.get("period") or {}
    subtitle = (
        f"Период: {period_days} дн. · "
        f"{str(period.get('start', ''))[:10]} — {str(period.get('end', ''))[:10]} · "
        f"сформировано {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    )
    row = _write_title(ws, 1, "Тригеры РА — рейтинг операторов (СП + VIP)", subtitle)

    headers = [
        "Оператор",
        "Обращений",
        "Неконструктив, %",
        "Δ неконструктив",
        "Негатив клиента, %",
        "Δ негатив",
        "Завершил оператор, %",
        "Δ завершение",
        "Оценка ОКК, %",
        "Δ ОКК",
        "В мониторинге, %",
        "Δ мониторинг",
        "Риск выгорания (ИИ)",
        "Δ риск ИИ",
        "Эмпатия (ИИ)",
        "Δ эмпатия",
        "Вовлечённость (ИИ)",
        "Преждевр. завершение, %",
        "Балл риска",
        "Зона риска",
        "Срабатывания",
    ]
    row = _write_headers(ws, row, headers)

    operators = data.get("operators") or []
    for i, op in enumerate(operators):
        display = op.get("display") or {}
        trends = op.get("trends") or {}
        risk = bool(op.get("isAtRisk"))
        alt = i % 2 == 1
        values = [
            op.get("operator") or "",
            op.get("totalCalls") or 0,
            display.get("nekonstruktivPct") or "—",
            (trends.get("nekonstruktiv") or {}).get("text") or "—",
            display.get("clientNegPct") or "—",
            (trends.get("clientNeg") or {}).get("text") or "—",
            display.get("operatorEndPct") or "—",
            (trends.get("operatorEnd") or {}).get("text") or "—",
            display.get("qcPct") or "—",
            (trends.get("qc") or {}).get("text") or "—",
            display.get("monitoringPct") or "—",
            (trends.get("monitoring") or {}).get("text") or "—",
            display.get("llmRisk") or "—",
            (trends.get("llmRisk") or {}).get("text") or "—",
            display.get("empathy") or "—",
            (trends.get("empathy") or {}).get("text") or "—",
            display.get("engagement") or "—",
            display.get("prematureClosurePct") or "—",
            display.get("score") or "—",
            "Да" if risk else "Нет",
            op.get("triggersLabel") or "—",
        ]
        trend_cols = {
            4: "nekonstruktiv",
            6: "clientNeg",
            8: "operatorEnd",
            10: "qc",
            12: "monitoring",
            14: "llmRisk",
            16: "empathy",
        }
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=val)
            _style_body_cell(cell, alt=alt, risk=risk)
            if col in trend_cols:
                sent = (trends.get(trend_cols[col]) or {}).get("sentiment") or "neutral"
                cell.font = _trend_font(str(sent))
            if col in (2, 19):
                cell.alignment = Alignment(horizontal="right", vertical="center")
            if col == 20 and risk:
                cell.font = Font(bold=True, color="CF222E", name="Calibri", size=10)
        row += 1

    _auto_width(ws, min_w=11, max_w=36)
    ws.freeze_panes = "B5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{max(4, row - 1)}"


def _sheet_at_risk(wb: Workbook, data: dict[str, Any]) -> None:
    ws = wb.create_sheet("Зона риска")
    row = _write_title(
        ws,
        1,
        "Операторы в зоне риска",
        "Балл ≥ 51 или ≥ 2 срабатываний одновременно",
    )
    headers = ["Оператор", "Обращений", "Балл риска", "Срабатывания"]
    row = _write_headers(ws, row, headers)
    for i, op in enumerate(data.get("atRisk") or []):
        display = op.get("display") or {}
        vals = [
            op.get("operator") or "",
            op.get("totalCalls") or 0,
            display.get("score") or "—",
            op.get("triggersLabel") or "—",
        ]
        for col, val in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=col, value=val)
            _style_body_cell(cell, alt=i % 2 == 1, risk=True)
        row += 1
    if not (data.get("atRisk") or []):
        ws.cell(row=row, column=1, value="Нет операторов в зоне риска").font = _SUBTITLE_FONT
    _auto_width(ws)


def _sheet_projects(wb: Workbook, data: dict[str, Any]) -> None:
    ws = wb.create_sheet("По проектам")
    row = _write_title(ws, 1, "Метрики по проектам СП / VIP", "")
    projects = data.get("projects") or []
    for project in projects:
        group = str(project.get("group") or "").upper()
        name = project.get("name") or project.get("id") or "Проект"
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        cell = ws.cell(row=row, column=1, value=f"{group} · {name}")
        cell.font = _SECTION_FONT
        cell.fill = _SECTION_FILL
        for c in range(1, 9):
            ws.cell(row=row, column=c).fill = _SECTION_FILL
            ws.cell(row=row, column=c).border = _THIN
        row += 1

        headers = [
            "Оператор",
            "Обращений",
            "Неконструктив, %",
            "Негатив клиента, %",
            "Завершил оператор, %",
            "Оценка ОКК, %",
            "В мониторинге, %",
            "Риск ИИ",
            "Эмпатия ИИ",
        ]
        row = _write_headers(ws, row, headers)
        for i, op in enumerate(project.get("operators") or []):
            qc = op.get("qcAvgWeight")
            vals = [
                op.get("operator") or "",
                op.get("totalCalls") or 0,
                _pct(op.get("nekonstruktivPct")),
                _pct(op.get("clientNegPct")) if project.get("hasClientSentiment") else "—",
                _pct(op.get("operatorEndPct")) if project.get("hasEndCall") else "—",
                _pct(qc * 100) if isinstance(qc, (int, float)) else "—",
                _pct(op.get("monitoringPct")),
                _pct(op.get("llmBurnoutAvg")),
                _pct(op.get("llmEmpathyAvg")),
            ]
            for col, val in enumerate(vals, start=1):
                cell = ws.cell(row=row, column=col, value=val)
                _style_body_cell(cell, alt=i % 2 == 1)
            row += 1
        row += 1
    _auto_width(ws)


def _sheet_tm(wb: Workbook, data: dict[str, Any]) -> None:
    ws = wb.create_sheet("ТМ Реактивация")
    row = _write_title(ws, 1, "ТМ · Реактивация", "")
    for section in data.get("tm") or []:
        name = section.get("name") or "ТМ"
        statuses: list[str] = list(section.get("statuses") or [])
        headers = [
            "Оператор",
            "Обращений",
            "Завершил оператор, %",
            "Негатив клиента, %",
            *statuses,
            "Риск ИИ",
            "Эмпатия ИИ",
            "Преждевр. завершение, %",
        ]
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=min(8, len(headers)))
        title_cell = ws.cell(row=row, column=1, value=name)
        title_cell.font = _SECTION_FONT
        title_cell.fill = _SECTION_FILL
        for c in range(1, len(headers) + 1):
            ws.cell(row=row, column=c).fill = _SECTION_FILL
            ws.cell(row=row, column=c).border = _THIN
        row += 1
        row = _write_headers(ws, row, headers)

        for i, op in enumerate(section.get("operators") or []):
            status_pcts = op.get("statusPcts") or {}
            vals: list[Any] = [
                op.get("operator") or "",
                op.get("totalCalls") or 0,
                _pct(op.get("operatorEndPct")),
                _pct(op.get("clientNegPct")),
            ]
            for st in statuses:
                vals.append(_pct(status_pcts.get(st)))
            vals.extend(
                [
                    _pct(op.get("llmBurnoutAvg")),
                    _pct(op.get("llmEmpathyAvg")),
                    _pct(op.get("llmPrematureClosurePct")),
                ]
            )
            for col, val in enumerate(vals, start=1):
                cell = ws.cell(row=row, column=col, value=val)
                _style_body_cell(cell, alt=i % 2 == 1)
            row += 1
        row += 1
    _auto_width(ws, max_w=28)


def _sheet_charts(wb: Workbook, data: dict[str, Any]) -> None:
    ws = wb.create_sheet("Динамика 30 дней")
    row = _write_title(ws, 1, "Графики 30 дней — дневные ряды", "")
    for chart in data.get("charts") or []:
        name = chart.get("name") or chart.get("id") or "Проект"
        group = str(chart.get("group") or "").upper()
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        cell = ws.cell(row=row, column=1, value=f"{group} · {name}")
        cell.font = _SECTION_FONT
        cell.fill = _SECTION_FILL
        for c in range(1, 4):
            ws.cell(row=row, column=c).fill = _SECTION_FILL
            ws.cell(row=row, column=c).border = _THIN
        row += 1
        headers = ["Дата", "Неконструктив, %", "Негатив клиента, %"]
        row = _write_headers(ws, row, headers)
        for i, point in enumerate(chart.get("points") or []):
            vals = [
                str(point.get("date") or "")[:10],
                _pct(point.get("nekonstruktivPct")),
                _pct(point.get("clientNegPct")) if chart.get("hasClientSentiment") else "—",
            ]
            for col, val in enumerate(vals, start=1):
                cell = ws.cell(row=row, column=col, value=val)
                _style_body_cell(cell, alt=i % 2 == 1)
            row += 1
        row += 1
    _auto_width(ws)


def build_triggers_ra_xlsx(data: dict[str, Any]) -> bytes:
    wb = Workbook()
    _sheet_operators(wb, data)
    _sheet_at_risk(wb, data)
    _sheet_projects(wb, data)
    _sheet_tm(wb, data)
    _sheet_charts(wb, data)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
